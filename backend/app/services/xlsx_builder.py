"""Сборка .xlsx из структуры DocumentStructure."""

from __future__ import annotations

import io
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.doc_gen_schema import DocumentStructure


def build_secretary_report_xlsx(
    records: list[dict],
    period_label: str,
) -> bytes:
    """
    Excel-отчёт агента «Учёт затрат»:
    - A–D: основная таблица (Дата / Категория / Сумма / Примечание)
    - F–G: итоги по категориям (Категория / Итого)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (period_label or "Отчёт")[:31]

    # ── Агрегация по категориям ───────────────────────────────────────────────
    totals: dict[str, int] = defaultdict(int)
    for r in records:
        cat = str(r.get("category") or "Прочие")
        try:
            totals[cat] += int(r.get("amount") or 0)
        except (TypeError, ValueError):
            pass
    cats_sorted = sorted(totals.items(), key=lambda x: -x[1])
    total_all = sum(v for _, v in cats_sorted) or 0

    # ── Стили ─────────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2E75B6")
    border_thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    def _hdr(row: int, col: int, val: str) -> None:
        c = ws.cell(row=row, column=col, value=val)
        c.font = header_font
        c.fill = header_fill
        c.border = border_thin
        c.alignment = Alignment(horizontal="center", vertical="center")

    def _cell(row: int, col: int, val, *, bold: bool = False, align: str = "left", num_format: str | None = None) -> None:
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, size=10)
        c.border = border_thin
        c.alignment = Alignment(horizontal=align, vertical="center")
        if num_format:
            c.number_format = num_format

    # ── 1. Таблица данных (A–D) ───────────────────────────────────────────────
    ROW_HDR = 1
    _hdr(ROW_HDR, 1, "Дата")
    _hdr(ROW_HDR, 2, "Категория")
    _hdr(ROW_HDR, 3, "Сумма")
    _hdr(ROW_HDR, 4, "Примечание")

    for i, r in enumerate(records, 2):
        raw_at = r.get("at", "")
        try:
            at_fmt = datetime.fromisoformat(str(raw_at).replace("Z", "+00:00")).strftime("%d.%m.%Y")
        except Exception:
            at_fmt = str(raw_at)[:10]
        _cell(i, 1, at_fmt, align="center")
        _cell(i, 2, str(r.get("category") or ""))
        try:
            amount_val = int(r.get("amount") or 0)
        except (TypeError, ValueError):
            amount_val = 0
        _cell(i, 3, amount_val, align="right", num_format="#,##0")
        _cell(i, 4, str(r.get("note") or r.get("description") or ""))

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 22

    # ── 2. Итоги по категориям (F–G) ─────────────────────────────────────────
    CATS_COL = 6   # F
    AMT_COL  = 7   # G
    _hdr(ROW_HDR, CATS_COL, "Категория")
    _hdr(ROW_HDR, AMT_COL,  "Итого")

    for i, (cat, total) in enumerate(cats_sorted, 2):
        _cell(i, CATS_COL, cat)
        _cell(i, AMT_COL, total, align="right", num_format="#,##0")

    # Строка «ИТОГО»
    if cats_sorted:
        total_row = len(cats_sorted) + 2
        c = ws.cell(row=total_row, column=CATS_COL, value="ИТОГО")
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.border = border_thin
        amt_c = ws.cell(row=total_row, column=AMT_COL, value=total_all)
        amt_c.font = Font(bold=True, size=10)
        amt_c.border = border_thin
        amt_c.number_format = "#,##0"
        amt_c.alignment = Alignment(horizontal="right")

    ws.column_dimensions[get_column_letter(CATS_COL)].width = 25
    ws.column_dimensions[get_column_letter(AMT_COL)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_report_xlsx_bytes(headers: list[str], rows: list[list[str]], sheet_name: str = "Отчёт") -> bytes:
    """
    Упрощённый builder для табличных отчётов — без заголовка документа и подписи таблицы.
    Первая строка = жирные заголовки столбцов, далее данные.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_xlsx_bytes(structure: DocumentStructure) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (structure.title or "Документ")[:31]

    row = 1
    title_cell = ws.cell(row=row, column=1, value=structure.title or "Документ")
    title_cell.font = Font(bold=True, size=14)
    row += 2

    for section in structure.sections:
        if section.heading:
            heading_cell = ws.cell(row=row, column=1, value=section.heading)
            heading_cell.font = Font(bold=True)
            row += 1
        for para in section.paragraphs:
            if para.strip():
                ws.cell(row=row, column=1, value=para.strip())
                row += 1
        row += 1

    for table_def in structure.tables:
        if table_def.caption:
            cap_cell = ws.cell(row=row, column=1, value=table_def.caption)
            cap_cell.font = Font(bold=True)
            row += 1
        col_count = max(len(table_def.headers), max((len(r) for r in table_def.rows), default=0))
        if col_count < 1:
            continue
        for col_idx in range(col_count):
            header = table_def.headers[col_idx] if col_idx < len(table_def.headers) else ""
            cell = ws.cell(row=row, column=col_idx + 1, value=header)
            cell.font = Font(bold=True)
        row += 1
        for table_row in table_def.rows:
            for col_idx in range(col_count):
                value = table_row[col_idx] if col_idx < len(table_row) else ""
                ws.cell(row=row, column=col_idx + 1, value=value)
            row += 1
        row += 2

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
