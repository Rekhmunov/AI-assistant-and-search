import csv
import io
import logging
from pathlib import Path

from app.constants.attachments import MAX_EXTRACT_CHARS_PER_FILE

logger = logging.getLogger(__name__)

MAX_EXTRACT_CHARS = MAX_EXTRACT_CHARS_PER_FILE

IMAGE_EXT = frozenset({"jpg", "jpeg", "png", "webp"})
DOCUMENT_EXT = frozenset({"txt", "md", "json", "csv", "pdf", "docx", "xlsx", "xls"})


def _clip(text: str) -> str:
    text = text.strip()
    if len(text) <= MAX_EXTRACT_CHARS:
        return text
    return text[:MAX_EXTRACT_CHARS] + "\n… [обрезано]"


def _decode_text_bytes(data: bytes) -> str:
    """UTF-8, BOM, Windows-1251 (типично для 1С/банк-клиент) и fallback."""
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        try:
            return data.decode("utf-16")
        except UnicodeDecodeError:
            pass
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def extract_text(filename: str, data: bytes) -> str:
    ext = Path(filename).suffix.lower().lstrip(".")
    if ext in ("txt", "md", "json"):
        return _clip(_decode_text_bytes(data))
    if ext == "csv":
        return _clip(_parse_csv(data))
    if ext == "pdf":
        return _clip(_parse_pdf(data))
    if ext == "docx":
        return _clip(_parse_docx(data))
    if ext == "xlsx":
        return _clip(_parse_xlsx(data))
    if ext == "xls":
        return _clip(_parse_xls(data))
    if ext in IMAGE_EXT:
        return _clip(_parse_image(data))
    raise ValueError(f"Unsupported extension: {ext}")


def is_image_filename(filename: str) -> bool:
    ext = Path(filename).suffix.lower().lstrip(".")
    return ext in IMAGE_EXT


def _parse_csv(data: bytes) -> str:
    text = _decode_text_bytes(data)
    reader = csv.reader(io.StringIO(text))
    rows = []
    for i, row in enumerate(reader):
        if i > 500:
            rows.append("…")
            break
        rows.append("\t".join(row))
    return "\n".join(rows)


def _parse_pdf(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    parts = []
    for page in reader.pages[:30]:
        parts.append(page.extract_text() or "")
    return "\n\n".join(parts)


def _parse_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _parse_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets[:5]:
        parts.append(f"## Лист: {sheet.title}")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i > 400:
                parts.append("…")
                break
            cells = ["" if c is None else str(c) for c in row]
            if any(cells):
                parts.append("\t".join(cells))
    wb.close()
    return "\n".join(parts)


def _parse_image(data: bytes) -> str:
    from PIL import Image
    import pytesseract

    try:
        img = Image.open(io.BytesIO(data))
    except Exception as e:
        raise ValueError("Не удалось открыть изображение") from e

    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")

    try:
        text = pytesseract.image_to_string(img, lang="rus+eng")
    except Exception:
        logger.exception("OCR failed")
        raise ValueError("Не удалось распознать текст на фото") from None

    if not text.strip():
        raise ValueError("На фото не найден текст. Попробуйте более чёткий снимок или загрузите PDF.")
    return text


def _parse_xls(data: bytes) -> str:
    import xlrd

    book = xlrd.open_workbook(file_contents=data, on_demand=True)
    parts = []
    for si in range(min(book.nsheets, 5)):
        sheet = book.sheet_by_index(si)
        parts.append(f"## Лист: {sheet.name}")
        for rx in range(min(sheet.nrows, 400)):
            row = [str(sheet.cell_value(rx, cx)) for cx in range(sheet.ncols)]
            if any(cell.strip() for cell in row):
                parts.append("\t".join(row))
    return "\n".join(parts)
